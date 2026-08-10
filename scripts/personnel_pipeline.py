#!/usr/bin/env python3
"""DAI/2 personnel pipeline.

Safe-by-default: the default audit path reads the existing data/data.js and never
writes to production files. The build path writes to an explicit output path and
requires a separate promotion step.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Any

ALLOWED_FIELDS = ("rank", "name", "org", "subunit", "phone")
FORBIDDEN_HINTS = {
    "cpf", "rg", "email", "e-mail", "address", "endereco", "endereço",
    "matricula", "matrícula", "numero bm", "número bm", "sexo", "funcao", "função"
}


def extract_assignment(text: str, variable: str) -> Any:
    pattern = rf"window\.{re.escape(variable)}\s*=\s*(.*?);(?:\r?\n|$)"
    match = re.search(pattern, text, flags=re.DOTALL)
    if not match:
        raise ValueError(f"Assignment window.{variable} not found")
    return json.loads(match.group(1))


def load_current(path: Path) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    text = path.read_text(encoding="utf-8")
    ddqod = extract_assignment(text, "DAI2_DDQOD")
    personnel = extract_assignment(text, "DAI2_PERSONNEL")
    if not isinstance(ddqod, dict) or not isinstance(personnel, list):
        raise ValueError("Unexpected data.js structure")
    return ddqod, personnel


def validate_personnel(ddqod: dict[str, Any], personnel: list[dict[str, Any]]) -> dict[str, Any]:
    errors: list[str] = []
    warnings: list[str] = []
    names: list[str] = []
    org_counts: Counter[str] = Counter()

    for index, row in enumerate(personnel, start=1):
        if not isinstance(row, dict):
            errors.append(f"row {index}: expected object")
            continue
        unexpected = sorted(set(row) - set(ALLOWED_FIELDS))
        if unexpected:
            errors.append(f"row {index}: non-allowlisted fields: {', '.join(unexpected)}")
        lowered = {str(k).lower() for k in row}
        leaked = sorted(h for h in FORBIDDEN_HINTS if h in lowered)
        if leaked:
            errors.append(f"row {index}: forbidden field hints: {', '.join(leaked)}")
        for field in ("rank", "name", "org"):
            if not str(row.get(field, "")).strip():
                errors.append(f"row {index}: required field '{field}' is empty")
        name = str(row.get("name", "")).strip()
        org = str(row.get("org", "")).strip()
        if name:
            names.append(name.casefold())
        if org:
            org_counts[org] += 1
            if org not in ddqod:
                errors.append(f"row {index}: unknown org '{org}'")
        phone = str(row.get("phone", "")).strip()
        if phone and len(re.sub(r"\D", "", phone)) < 8:
            warnings.append(f"row {index}: phone appears unusually short")

    duplicates = sorted(name for name, count in Counter(names).items() if count > 1)
    if duplicates:
        errors.append("duplicate personnel names detected: " + ", ".join(duplicates))

    planned_by_org = {
        org: sum(int(v or 0) for v in grades.values())
        for org, grades in ddqod.items()
        if isinstance(grades, dict)
    }
    capacity_warnings = []
    for org, existing in sorted(org_counts.items()):
        planned = planned_by_org.get(org)
        if planned is not None and existing > planned:
            capacity_warnings.append(f"{org}: existing {existing} exceeds DDQOD {planned}")
    warnings.extend(capacity_warnings)

    return {
        "ok": not errors,
        "records": len(personnel),
        "organizations": dict(sorted(org_counts.items())),
        "planned_positions": sum(planned_by_org.values()),
        "errors": errors,
        "warnings": warnings,
    }


def audit(args: argparse.Namespace) -> int:
    ddqod, personnel = load_current(Path(args.input))
    report = validate_personnel(ddqod, personnel)
    payload = json.dumps(report, ensure_ascii=False, indent=2)
    print(payload)
    if args.report:
        out = Path(args.report)
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(payload + "\n", encoding="utf-8")
    return 0 if report["ok"] else 1


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def build_from_xlsx(args: argparse.Namespace) -> int:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("openpyxl is required for XLSX build mode") from exc

    baseline_ddqod, baseline_personnel = load_current(Path(args.baseline))
    workbook = load_workbook(args.xlsx, read_only=True, data_only=True)
    sheet = workbook[args.sheet] if args.sheet else workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [normalize_header(v) for v in next(rows)]

    aliases = {
        "rank": {"posto/graduação", "posto/graducao", "p/g", "posto", "graduação", "graduacao"},
        "name": {"nome", "nome completo"},
        "org": {"órgão", "orgao", "órgão externo", "orgao externo"},
        "subunit": {"unidade", "subunidade", "ctpm", "unidade ctpm"},
        "phone": {"contato", "telefone", "celular"},
    }
    index_by_field: dict[str, int] = {}
    for field, names in aliases.items():
        for idx, header in enumerate(headers):
            if header in names:
                index_by_field[field] = idx
                break
    missing = [f for f in ("rank", "name", "org") if f not in index_by_field]
    if missing:
        raise SystemExit("missing required XLSX columns: " + ", ".join(missing))

    # Privacy gate: reject known sensitive columns instead of silently carrying them forward.
    sensitive_headers = [h for h in headers if any(token in h for token in FORBIDDEN_HINTS)]
    if sensitive_headers and not args.allow_sensitive_source:
        raise SystemExit(
            "source contains sensitive columns; inspect and rerun with --allow-sensitive-source only "
            "after confirming they will not be emitted: " + ", ".join(sensitive_headers)
        )

    generated: list[dict[str, str]] = []
    for values in rows:
        name_idx = index_by_field["name"]
        if name_idx >= len(values) or not str(values[name_idx] or "").strip():
            continue
        item = {}
        for field in ALLOWED_FIELDS:
            idx = index_by_field.get(field)
            item[field] = str(values[idx] or "").strip() if idx is not None and idx < len(values) else ""
        generated.append(item)

    report = validate_personnel(baseline_ddqod, generated)
    if not report["ok"]:
        print(json.dumps(report, ensure_ascii=False, indent=2))
        return 1

    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        "window.DAI2_DDQOD = " + json.dumps(baseline_ddqod, ensure_ascii=False, separators=(",", ":")) + ";\n"
        + "window.DAI2_PERSONNEL = " + json.dumps(generated, ensure_ascii=False, separators=(",", ":")) + ";\n",
        encoding="utf-8",
    )

    old = {r.get("name", "").casefold(): r for r in baseline_personnel}
    new = {r.get("name", "").casefold(): r for r in generated}
    diff = {
        "baseline_records": len(baseline_personnel),
        "generated_records": len(generated),
        "added": [new[k] for k in sorted(new.keys() - old.keys())],
        "removed": [old[k] for k in sorted(old.keys() - new.keys())],
        "changed": [
            {"before": old[k], "after": new[k]}
            for k in sorted(old.keys() & new.keys()) if old[k] != new[k]
        ],
        "validation": report,
    }
    diff_path = Path(args.diff_report)
    diff_path.parent.mkdir(parents=True, exist_ok=True)
    diff_path.write_text(json.dumps(diff, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"generated {output}; review {diff_path} before promotion")
    return 0


def parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="DAI/2 personnel safety pipeline")
    sub = p.add_subparsers(dest="command", required=True)
    a = sub.add_parser("audit", help="audit current data.js without changing it")
    a.add_argument("--input", default="data/data.js")
    a.add_argument("--report", default="")
    a.set_defaults(func=audit)

    b = sub.add_parser("build", help="generate a candidate data.js from XLSX")
    b.add_argument("--xlsx", required=True)
    b.add_argument("--sheet", default="")
    b.add_argument("--baseline", default="data/data.js")
    b.add_argument("--output", default="artifacts/data.generated.js")
    b.add_argument("--diff-report", default="artifacts/personnel-diff.json")
    b.add_argument("--allow-sensitive-source", action="store_true")
    b.set_defaults(func=build_from_xlsx)
    return p


def main() -> int:
    args = parser().parse_args()
    return int(args.func(args))


if __name__ == "__main__":
    sys.exit(main())
